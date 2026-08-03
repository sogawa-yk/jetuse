"""xlsx の前処理: シート名とセル範囲つきのチャンク抽出(PREP-01 / ADR-0020)。

「仕様書 v2.0 の『制約』シート C5:E5 が根拠」まで返せるようにするための**入口**。
メタデータの配管(`rag_metadata.ATTRIBUTE_KEYS` / `rag_adb` のチャンク単位の列)は
RAGM-01 / RAGM-02 で完成しているので、ここは抽出だけを担う。

出典の粒度はバックエンドで**違う**(ADR-0020 の決定内容そのもの):

| バックエンド | xlsx の出典粒度 |
|---|---|
| `adb` | **チャンク単位**。シートとセル範囲がチャンクごとに異なる |
| `vector_store`(マネージド) | **ファイル単位**。属性は 1 ファイルに 1 種類(SPIKE-M1 ①-a 実測) |

そのため `file_attributes()` はファイル全体を表す値を返す。**チャンクごとの sheet / cells を
マネージド側の属性に載せることはできない**(1 チャンク = 1 ファイルに割って
「セル単位で返る」ように見せる細工はしない — 能力差を隠さないこと自体が決定事項)。

読み方は `read_only=True, data_only=True`:
- `read_only`: 大きなブックを一度にメモリへ展開しない(行を逐次読む)。
- `data_only`: 数式ではなく**値**(Excel が保存したキャッシュ)を取る。キャッシュが無い
  数式セル(openpyxl で書いたブック等)は `None` = 空セル扱いになる。

上限を超えたら**切り詰めずに拒否する**(`ExtractionLimitError` → ルート側で 422)。
黙って一部だけ取り込むと、利用者からは「取り込めたのに根拠が出ない」と区別が付かない
(RAGM-01 で属性の切り詰めを禁じたのと同じ理由)。

ただし**1 チャンクの文字数**(`chunk_chars`)は拒否の理由にしない(PREP-04)。分ける場所が
無いなら作る — 行境界 → セル境界 → **セルの中**の順に割り、どの断片も元のセルを出典に持つ。
拒否していた頃は「1 セルに巨大なテキスト」というだけでファイル全体が入らなかった
(実案件で 8 冊中 2 冊)。総チャンク数・バイト数の上限は据え置きで、超えれば従来どおり 422。
"""

import contextlib
import datetime
import io
import logging
import re
import zipfile
from typing import Any
from xml.etree import ElementTree

logger = logging.getLogger("jetuse.extract_xlsx")

# 上限(超過は切り詰めずに拒否する)。
# xlsx は zip なので、生バイト数より展開後(共有文字列 + シート XML)が桁で大きい。
# 汎用アップロード上限(`rag.MAX_BYTES` = 20MB)より手前で止める。
MAX_WORKBOOK_BYTES = 10 * 1024 * 1024
# 展開後(zip 内の各エントリの合計)の上限。**圧縮率で殴られないため**に、開く前に見る。
# 10MB の zip は数百 MB に展開されうる(zip bomb)。
MAX_UNCOMPRESSED_BYTES = 100 * 1024 * 1024
# 走査してよいセル数の上限。上のバイト上限を通っても、**中身が無いのに広いシート**は
# 「チャンクにならないまま延々と読む」ことができてしまう(`read_only` の反復は欠けた行を
# 空行で埋めるので、`A1` と `A1048576` だけの数 KB のブックが 100 万行の反復になる。
# 列も同じで、1 行あたり最大 16,384 セルまで膨らむ)。**非空行ではなく反復したセル**を
# 数える — 空行しか無い領域では非空行の数え上げが 1 つも増えず、上限として働かない。
MAX_SCANNED_CELLS = 5_000_000
MAX_CHUNKS = 1_000
# 1 チャンクの文字数。埋め込み API は先頭 2,000 文字で切るので、保存本文もそこで揃える
# (超える塊を作ると、本文と検索表現が食い違う)。
MAX_CHUNK_CHARS = 2_000
# 1 セルが上限を超えたときに、セルの中を切ってよい位置(PREP-04)。改行が最優先で、
# 無ければこれらの**直後**で切る。無ければ文字数で切る。構造の解析はしない(非ゴール)。
_SEPARATORS = ("。", "、", "，", "；", "！", "？", ".", ",", ";", " ", "\t")

# ファイル単位の属性(マネージド Vector Store 用)で「ブック全体」を表す値。
WORKBOOK_SHEET = "(ブック全体: {n} シート)"
WORKBOOK_CELLS = "(ブック全体)"

XLSX_EXT = ".xlsx"

_A1 = re.compile(r"^([A-Z]+)(\d+)$")


def is_xlsx(filename: str) -> bool:
    """取り込み経路の分岐点(拡張子の判定はここ 1 か所に置く)。"""
    return (filename or "").lower().endswith(XLSX_EXT)


class XlsxExtractError(ValueError):
    """xlsx を抽出できない。ルート側で 422 に正規化する。"""


class UnsupportedWorkbook(XlsxExtractError):
    """xlsx として開けない(壊れている / 別形式 / 暗号化)。"""


class EmptyWorkbook(XlsxExtractError):
    """開けたが本文が 1 文字も無い。取り込み経路では拒否する(空の本文を送らない)。"""


class ExtractionLimitError(XlsxExtractError):
    """上限超過。**どの上限か**を `limit` と本文の両方に持つ(422 の detail に出す)。"""

    def __init__(self, limit: str, message: str):
        super().__init__(f"{message} (limit={limit})")
        self.limit = limit


def extract(filename: str, content: bytes) -> list[dict[str, Any]]:
    """xlsx を `{sheet, cells, text}` のチャンク列にする。

    シートごとに**連続する非空セルの矩形**(空行 / 行の飛びで区切られる塊)を 1 チャンクに
    まとめ、`sheet`(シート名)と `cells`(A1 形式。例 `B12:F48`)を付ける。
    1 チャンクが上限文字数に収まらないときは**行境界で分割**する(切り詰めない。
    分割した各チャンクは自分の行範囲を `cells` に持つ)。1 行だけで超えるときは
    **セル境界**で、1 セルだけで超えるときは**セルの中**で割る(PREP-04)。セル内で
    割った断片は `part`(`2/7`)を持ち、`cells` は元のセルのまま。

    **行もチャンクも逐次処理し、上限は 1 件作るたびに見る**。塊を丸ごとメモリに溜めてから
    数えると、上限内のファイルでも(密なシート 1 枚で数十万行)worker を落とせる。
    同時に保持するのは「作りかけの 1 チャンク分の行」と、確定したチャンク列だけ
    (どちらも上限で抑えてある)。
    """
    if len(content) > MAX_WORKBOOK_BYTES:
        raise ExtractionLimitError(
            "workbook_bytes",
            f"ブックのバイト数が上限を超えました({len(content)} > {MAX_WORKBOOK_BYTES} バイト)",
        )
    _check_uncompressed(content)
    workbook = _open(content)
    chunks: list[dict[str, Any]] = []
    try:
        with _normalised_errors():
            budget = _Budget()
            for sheet in workbook.worksheets:
                for chunk in _sheet_chunks(sheet, budget):
                    chunks.append(chunk)
                    if len(chunks) > MAX_CHUNKS:
                        # 全部展開しきってから数えない(上限を超えた時点で止める)
                        raise ExtractionLimitError(
                            "chunks",
                            f"チャンク数が上限を超えました(> {MAX_CHUNKS})。"
                            "シートを分けるかファイルを分割してください",
                        )
    finally:
        workbook.close()
    return chunks


def file_attributes(chunks: list[dict[str, Any]]) -> dict[str, str]:
    """マネージド Vector Store へ渡す**ファイル単位**の属性(`sheet` / `cells`)。

    属性はファイル単位でしか保持されない(SPIKE-M1 ①-a)。1 ファイルが複数チャンクに
    割れても属性は 1 種類なので、ここが返せるのは「そのファイル全体の範囲」だけである。
    単一シートなら実際の外接範囲、複数シートなら「ブック全体」であることを明示する。
    """
    if not chunks:
        return {}
    sheets = list(dict.fromkeys(c["sheet"] for c in chunks))
    if len(sheets) > 1:
        return {"sheet": WORKBOOK_SHEET.format(n=len(sheets)), "cells": WORKBOOK_CELLS}
    bounds = [_parse_range(c["cells"]) for c in chunks]
    return {
        "sheet": sheets[0],
        "cells": _a1_range(
            min(b[0] for b in bounds), min(b[1] for b in bounds),
            max(b[2] for b in bounds), max(b[3] for b in bounds),
        ),
    }


def render_text(chunks: list[dict[str, Any]]) -> str:
    """マネージド Vector Store へ渡す本文(テキスト化)。

    マネージド側は xlsx を受け付けない(SPIKE-03 で docx が `Unsupported file type`。
    xlsx も同じ扱いであることは PREP-01 の E2E で実測する)ため、抽出結果をテキストとして
    渡す。見出しに出典を書くが、**属性はファイル単位のまま**であり、これでチャンクごとの
    属性が持てるようになるわけではない。
    """
    return "\n\n".join(f"[{c['sheet']} {c['cells']}]\n{c['text']}" for c in chunks)


# --- 読み取り -----------------------------------------------------------------

# openpyxl は「壊れたブック」を複数の型で投げてくる(zip 層 / XML 層 / 内部構造)。
# 500 として漏らすと、壊れたファイル 1 つでルートの契約(不正な入力は 422)が破れる。
# `RuntimeError` は**暗号化 zip**（パスワード付きブック）で `zipfile` が投げる型。
# ValueError 系ではないので、入れ忘れると利用者入力で 500 が漏れる(review-5 PREP01-017)。
_BROKEN = (zipfile.BadZipFile, ElementTree.ParseError, KeyError, ValueError, TypeError,
           AttributeError, IndexError, OSError, RuntimeError)


@contextlib.contextmanager
def _normalised_errors():
    """壊れたブック由来の例外を `UnsupportedWorkbook` に寄せる(上限超過は素通し)。

    **開くときだけでなく、行を読んでいる最中も**必要。`read_only` の反復は遅延実行なので、
    シート XML の破損は `iter_rows()` の途中で初めて現れる。
    """
    try:
        yield
    except XlsxExtractError:
        raise  # 上限超過・空ブックは「壊れている」ではない
    except _BROKEN as e:
        raise UnsupportedWorkbook(f"xlsx として読めませんでした: {type(e).__name__}") from e


def _check_uncompressed(content: bytes) -> None:
    """展開後の合計サイズを**開く前に**見る(圧縮率で殴られないため)。"""
    with _normalised_errors(), zipfile.ZipFile(io.BytesIO(content)) as archive:
        total = sum(info.file_size for info in archive.infolist())
    if total > MAX_UNCOMPRESSED_BYTES:
        raise ExtractionLimitError(
            "uncompressed_bytes",
            f"展開後のサイズが上限を超えました({total} > {MAX_UNCOMPRESSED_BYTES} バイト)",
        )


class _Budget:
    """1 回の抽出で走査してよいセル数の残り(尽きたら上限として拒否する)。"""

    def __init__(self) -> None:
        self.cells = MAX_SCANNED_CELLS

    def spend(self, cells: int) -> None:
        self.cells -= max(cells, 1)  # 空行(0 セル)も反復コストは掛かる
        if self.cells < 0:
            raise ExtractionLimitError(
                "scanned_cells",
                f"走査したセル数が上限を超えました(> {MAX_SCANNED_CELLS})。"
                "使っていない広大な領域を含むブックはシートを整理してください",
            )


def _open(content: bytes):
    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException

    try:
        with _normalised_errors():
            return load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except InvalidFileException as e:
        raise UnsupportedWorkbook(f"xlsx として読めませんでした: {type(e).__name__}") from e


def _cell_text(value: Any) -> str:
    """セルの値を 1 つの文字列にする(表構造の復元はしない)。"""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float):
        # Excel の数値は float で返る。5.0 を "5.0" と書くと表として読みにくい
        return str(int(value)) if value.is_integer() else str(value)
    if isinstance(value, datetime.datetime):
        # Excel の日付は時刻つきの datetime で返る。0 時ちょうどは日付だけを書く
        # (「2026-07-30」を「2026-07-30T00:00:00」と書くと表として読みにくい)
        return value.date().isoformat() if value.time() == datetime.time() else value.isoformat()
    if isinstance(value, datetime.date | datetime.time):
        return value.isoformat()
    return str(value).strip()


def _column_index(cell: Any) -> int | None:
    """セルの列番号。`read_only` の穴埋め(`EmptyCell`)は座標を持たないので None。"""
    col = getattr(cell, "column", None)
    if isinstance(col, int):
        return col
    if isinstance(col, str):
        from openpyxl.utils import column_index_from_string

        return column_index_from_string(col)
    return None


def _rows(sheet, budget: _Budget):
    """`(行番号, {列番号: 文字列})` を**非空行だけ**返す(走査量は空行も含めて数える)。

    行番号は必ずセル自身(`cell.row`)から取る。`read_only` では空行が詰められることが
    あり、列挙の順番を行番号と見なすと出典がずれる。
    """
    try:
        iterator = sheet.iter_rows()
    except ValueError:
        # 空シートは寸法を決められずに ValueError になることがある(read_only)
        return
    for row in iterator:
        budget.spend(len(row))  # **反復した分**を数える(空行を飛ばして数えない)
        values: dict[int, str] = {}
        index: int | None = None
        for cell in row:
            col = _column_index(cell)
            if col is None:
                continue
            text = _cell_text(cell.value)
            if not text:
                continue  # 結合セルの 2 つ目以降・キャッシュの無い数式セルもここに落ちる
            values[col] = text
            if index is None:
                index = cell.row
        if values and index is not None:
            yield index, values


def _line(min_col: int, values: dict[int, str]) -> str:
    """1 行をタブ区切りにする(チャンクの左端から埋めるので列がずれない)。"""
    return "\t".join(values.get(col, "") for col in range(min_col, max(values) + 1))


def _sheet_chunks(sheet, budget: _Budget):
    """1 シートを**逐次**チャンクにする(塊全体をメモリに溜めない)。

    区切りは 2 つ: (a) 空行 / 行の飛び = 別の矩形、(b) 上限文字数 = 行境界で分割。
    保持するのは作りかけの 1 チャンク分の行だけ（上限文字数で抑えられる）。
    """
    buffer: list[tuple[int, dict[int, str]]] = []
    min_col: int | None = None
    size = 0
    previous: int | None = None
    for index, values in _rows(sheet, budget):
        if buffer and previous is not None and index != previous + 1:
            yield _chunk(sheet.title, buffer, min_col)
            buffer, min_col, size = [], None, 0
        row_min = min(values)
        if buffer and min_col is not None and row_min < min_col:
            # 左に広がった: 既存行の描画幅が変わるので測り直す。**広げた結果、既存分だけで
            # 上限を超えるなら、既存は元の左端のまま確定する**(広げたまま出すと上限超えの
            # チャンクができ、埋め込みの切り詰め位置と保存本文がずれる)
            widened = sum(len(_line(row_min, v)) for _, v in buffer) + len(buffer) - 1
            if widened > MAX_CHUNK_CHARS:
                yield _chunk(sheet.title, buffer, min_col)
                buffer, min_col, size = [], None, 0
            else:
                min_col, size = row_min, widened
        new_min = row_min if not buffer else min_col
        line_length = len(_line(new_min, values))
        if line_length > MAX_CHUNK_CHARS:
            # 1 行だけで上限を超える(PREP-04)。**セル境界で割り、それでも収まらないセルは
            # その中で割る**。拒否するとファイル全体が入らず、読み飛ばすと黙って欠ける
            if buffer:
                yield _chunk(sheet.title, buffer, min_col)
                buffer, min_col, size = [], None, 0
            yield from _long_row_chunks(sheet.title, index, values)
            previous = index
            continue
        if buffer and size + 1 + line_length > MAX_CHUNK_CHARS:
            yield _chunk(sheet.title, buffer, min_col)
            buffer, min_col, size = [], None, 0
            new_min = row_min
            line_length = len(_line(new_min, values))
        buffer.append((index, values))
        min_col = new_min
        size += line_length + (1 if size else 0)
        previous = index
    if buffer:
        yield _chunk(sheet.title, buffer, min_col)


def _long_row_chunks(sheet: str, row: int, values: dict[int, str]):
    """上限を超える 1 行を**セル境界**で割る。1 セルで超えるならそのセルの中で割る。

    実データは「非空セル 1 個・13,728 文字」だったので、セル境界だけでは割れない
    (PREP-04)。表構造の復元はしない — 割った断片は元のセルを出典に持つだけ。
    """
    group: list[int] = []          # 同じチャンクに入れる列
    size = 0
    for col in sorted(values):
        text = values[col]
        if len(text) > MAX_CHUNK_CHARS:
            if group:
                yield _row_chunk(sheet, row, group, values)
                group, size = [], 0
            yield from _cell_chunks(sheet, row, col, text)
            continue
        # `_line` は間の空セルもタブで埋めるので、列の飛びの分を足して測る
        added = len(text) + (col - group[-1] if group else 0)
        if group and size + added > MAX_CHUNK_CHARS:
            yield _row_chunk(sheet, row, group, values)
            group, size = [], 0
            added = len(text)
        group.append(col)
        size += added
    if group:
        yield _row_chunk(sheet, row, group, values)


def _row_chunk(sheet: str, row: int, group: list[int], values: dict[int, str]) -> dict:
    """行の一部(連続した列の並び)を 1 チャンクにする。出典はその列範囲。"""
    subset = {col: values[col] for col in group}
    return {
        "sheet": sheet,
        "cells": _a1_range(group[0], row, group[-1], row),
        "text": _line(group[0], subset),
    }


def _cell_chunks(sheet: str, row: int, col: int, text: str):
    """1 セルの中を割る。**どの断片も出典は元のセル**(そのセルが根拠なので精度は落ちない)。

    分割したことは `part`(`2/7`)で分かるようにする(黙って分割して終わりにしない)。
    既存の鍵は増減しないので、`part` を見ない取り込み経路は従来どおり動く。
    """
    cells = _a1_range(col, row, col, row)
    pieces = _split_text(text)
    logger.warning(
        "extract_xlsx: 1 セルが上限を超えたのでセル内で分割しました "
        "(sheet=%r cells=%s chars=%d parts=%d)", sheet, cells, len(text), len(pieces),
    )
    for i, piece in enumerate(pieces, start=1):
        yield {"sheet": sheet, "cells": cells, "text": piece,
               "part": f"{i}/{len(pieces)}"}


def _split_text(text: str) -> list[str]:
    """上限に収まる断片へ切る。**意味の切れ目を優先**(改行 → 区切り文字 → 文字数)。

    断片をつなぐと元の文字列に戻る(切り詰めない)。構造の解析はしない — JSON として
    妥当な位置で切ることは目指さない(非ゴール)。
    """
    pieces: list[str] = []
    start = 0
    while len(text) - start > MAX_CHUNK_CHARS:
        window = text[start:start + MAX_CHUNK_CHARS]
        cut = window.rfind("\n") + 1
        if cut <= 0:
            # 区切り文字は**その直後**で切る(区切り自体は前の断片に残す)
            cut = max(window.rfind(s) + len(s) for s in _SEPARATORS)
        if cut <= 0:
            cut = MAX_CHUNK_CHARS      # 切れ目が無い塊は文字数で切る(最後の手段)
        pieces.append(text[start:start + cut])
        start += cut
    pieces.append(text[start:])
    return pieces


def _chunk(sheet: str, buffer: list[tuple[int, dict[int, str]]], min_col: int | None) -> dict:
    left = min_col if min_col is not None else min(min(values) for _, values in buffer)
    max_col = max(max(values) for _, values in buffer)
    return {
        "sheet": sheet,
        "cells": _a1_range(left, buffer[0][0], max_col, buffer[-1][0]),
        "text": "\n".join(_line(left, values) for _, values in buffer),
    }


# --- A1 表記 ------------------------------------------------------------------


def _a1_range(min_col: int, min_row: int, max_col: int, max_row: int) -> str:
    from openpyxl.utils import get_column_letter

    start = f"{get_column_letter(min_col)}{min_row}"
    end = f"{get_column_letter(max_col)}{max_row}"
    return start if start == end else f"{start}:{end}"


def _parse_range(cells: str) -> tuple[int, int, int, int]:
    """`B12:F48` / `C5` を `(min_col, min_row, max_col, max_row)` に戻す。"""
    from openpyxl.utils import column_index_from_string

    parts = cells.split(":")
    bounds = []
    for part in parts:
        m = _A1.match(part.strip().upper())
        if not m:
            raise ValueError(f"A1 形式として読めない範囲: {cells}")
        bounds.append((column_index_from_string(m.group(1)), int(m.group(2))))
    (c1, r1), (c2, r2) = (bounds[0], bounds[-1])
    return (min(c1, c2), min(r1, r2), max(c1, c2), max(r1, r2))
