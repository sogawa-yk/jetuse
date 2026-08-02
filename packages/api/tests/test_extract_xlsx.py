"""xlsx 抽出(PREP-01)の単体テスト。

固定したいのは 4 つ:
1. **チャンクごとに異なる sheet / cells** が付く(adb バックエンドの出典粒度)。
2. マネージド Vector Store 向けの属性は**ファイル単位**である(チャンク単位に見せない)。
3. 上限超過は**切り詰めずに拒否**し、どの上限かが分かる。
4. 空シート・結合セル・数式セル・巨大な空白領域で落ちない。
"""

import io

import pytest
from openpyxl import Workbook

from jetuse_core import extract_xlsx


def build(sheets: dict[str, list[tuple[str, object]]], *, merge: dict[str, str] | None = None,
          formula: dict[str, str] | None = None) -> bytes:
    """架空のブックを作る(セル参照 → 値の指定)。顧客データは使わない。"""
    wb = Workbook()
    wb.remove(wb.active)
    for title, cells in sheets.items():
        ws = wb.create_sheet(title)
        for ref, value in cells:
            ws[ref] = value
        if merge and title in merge:
            ws.merge_cells(merge[title])
        if formula and title in formula:
            ws[formula[title]] = "=SUM(B1:B9)"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


SPEC = {
    "API一覧": [
        ("B12", "エンドポイント"), ("C12", "メソッド"), ("D12", "説明"),
        ("B13", "/v1/inventory"), ("C13", "GET"), ("D13", "在庫照会"),
    ],
    "制約": [
        ("C5", "レート制限"), ("D5", "600 req/min"), ("E5", "超過時は429"),
    ],
    "空シート": [],
}


def test_chunks_carry_sheet_and_cells():
    chunks = extract_xlsx.extract("サンプル仕様書.xlsx", build(SPEC))
    by_sheet = {c["sheet"]: c for c in chunks}
    assert set(by_sheet) == {"API一覧", "制約"}          # 空シートはチャンクを作らない
    assert by_sheet["API一覧"]["cells"] == "B12:D13"
    assert by_sheet["制約"]["cells"] == "C5:E5"
    assert "600 req/min" in by_sheet["制約"]["text"]
    assert by_sheet["API一覧"]["text"].splitlines()[0] == "エンドポイント\tメソッド\t説明"
    # チャンクごとに出典が異なる(= adb バックエンドで「どのセルが根拠か」を返せる)
    assert len({c["cells"] for c in chunks}) == len(chunks)


def test_blank_rows_split_blocks_and_keep_row_numbers():
    content = build({"S": [("A1", "見出し"), ("A2", "本文"), ("A10", "別の塊"), ("B10", "値")]})
    chunks = extract_xlsx.extract("a.xlsx", content)
    assert [c["cells"] for c in chunks] == ["A1:A2", "A10:B10"]


def test_empty_workbook_yields_no_chunks():
    assert extract_xlsx.extract("empty.xlsx", build({"空1": [], "空2": []})) == []


def test_merged_cells_do_not_crash():
    content = build({"S": [("A1", "結合された見出し"), ("A2", "値")]}, merge={"S": "A1:C1"})
    chunks = extract_xlsx.extract("merged.xlsx", content)
    assert len(chunks) == 1
    # 結合セルは左上だけが値を持つ(意味の解釈はしない = 非ゴール)
    assert chunks[0]["cells"] == "A1:A2"
    assert "結合された見出し" in chunks[0]["text"]


def test_formula_without_cached_value_is_treated_as_empty():
    """data_only=True では、キャッシュの無い数式セルは None で返る(落ちないこと)。"""
    content = build({"S": [("A1", "合計"), ("B1", 10), ("B2", 20)]}, formula={"S": "B3"})
    chunks = extract_xlsx.extract("formula.xlsx", content)
    assert len(chunks) == 1
    assert "=SUM" not in chunks[0]["text"]        # 数式そのものは取り込まない
    assert chunks[0]["cells"] == "A1:B2"


def test_large_blank_region_does_not_crash():
    content = build({"S": [("A1", "先頭"), ("A5000", "末尾")]})
    chunks = extract_xlsx.extract("sparse.xlsx", content)
    assert [c["cells"] for c in chunks] == ["A1", "A5000"]


def test_values_are_normalised():
    import datetime

    content = build({"S": [
        ("A1", 5.0), ("B1", 1.5), ("C1", True), ("D1", datetime.date(2026, 7, 30)),
        ("E1", "  余白つき  "),
    ]})
    assert extract_xlsx.extract("v.xlsx", content)[0]["text"] == (
        "5\t1.5\tTRUE\t2026-07-30\t余白つき"
    )


# --- 上限(切り詰めずに拒否する) ----------------------------------------------


def test_workbook_bytes_limit_is_rejected(monkeypatch):
    monkeypatch.setattr(extract_xlsx, "MAX_WORKBOOK_BYTES", 10)
    with pytest.raises(extract_xlsx.ExtractionLimitError) as e:
        extract_xlsx.extract("big.xlsx", build(SPEC))
    assert e.value.limit == "workbook_bytes"
    assert "limit=workbook_bytes" in str(e.value)


def test_chunk_count_limit_is_rejected(monkeypatch):
    monkeypatch.setattr(extract_xlsx, "MAX_CHUNKS", 2)
    rows = [(f"A{n}", f"値{n}") for n in range(1, 20, 2)]  # 空行で区切られた 10 個の塊
    with pytest.raises(extract_xlsx.ExtractionLimitError) as e:
        extract_xlsx.extract("many.xlsx", build({"S": rows}))
    assert e.value.limit == "chunks"


# --- 1 セルが上限を超える(PREP-04: セルの中で分割する) ------------------------


def _giant_cell_text(length: int) -> str:
    """**架空**の「1 セルに丸ごと入った API リクエスト例」(実データと同じ形)。"""
    block = ('{\n  "endpoint": "/v1/inventory",\n  "method": "POST",\n'
             '  "body": {"sku": "SKU-0001", "qty": 10}\n}\n')
    return (block * (length // len(block) + 1))[:length].strip()


def test_single_cell_over_limit_is_split_inside_the_cell():
    """1 セルに 13,000 文字級。**セル境界でも行境界でも割れない**ので、セルの中で割る。

    実案件では 13,728 文字 / 非空セル 1 個という形で、ファイル全体が 422 になっていた。
    """
    text = _giant_cell_text(13_000)
    chunks = extract_xlsx.extract("spec.xlsx", build({"サンプル": [("A53", text)]}))
    assert len(chunks) > 1
    assert all(len(c["text"]) <= extract_xlsx.MAX_CHUNK_CHARS for c in chunks)
    # 断片はどれも「そのセルが根拠」= 出典は元のセルのまま(精度は落ちない)
    assert {(c["sheet"], c["cells"]) for c in chunks} == {("サンプル", "A53")}
    assert "".join(c["text"] for c in chunks) == text        # 1 文字も落ちていない


def test_split_fragments_are_marked_so_the_split_is_not_silent():
    """黙って分割しない。断片には `part`(何分割の何番目か)が付く。分割していなければ無い。"""
    chunks = extract_xlsx.extract("spec.xlsx",
                                  build({"S": [("A1", _giant_cell_text(5_000))]}))
    n = len(chunks)
    assert [c["part"] for c in chunks] == [f"{i}/{n}" for i in range(1, n + 1)]
    assert all("part" not in c for c in extract_xlsx.extract("spec.xlsx", build(SPEC)))


def test_long_cell_is_split_at_newline_boundaries(monkeypatch):
    """切る位置は**意味の切れ目**が先(改行)。文字数で機械的に切らない。"""
    monkeypatch.setattr(extract_xlsx, "MAX_CHUNK_CHARS", 50)
    lines = [f'"field{n}": "値{n}",' for n in range(1, 30)]
    chunks = extract_xlsx.extract("spec.xlsx", build({"S": [("A1", "\n".join(lines))]}))
    assert len(chunks) > 1
    for c in chunks:
        assert all(line in lines for line in c["text"].splitlines())   # 行の途中で切らない


def test_long_cell_without_newlines_falls_back_to_separators(monkeypatch):
    """改行が無ければ区切り文字で切る(最後の手段が文字数)。"""
    monkeypatch.setattr(extract_xlsx, "MAX_CHUNK_CHARS", 50)
    text = "".join(f"項目{n}は必須である。" for n in range(1, 20))
    chunks = extract_xlsx.extract("spec.xlsx", build({"S": [("A1", text)]}))
    assert len(chunks) > 1
    assert all(c["text"].endswith("。") for c in chunks)
    assert "".join(c["text"] for c in chunks) == text


def test_long_cell_without_any_boundary_is_split_by_length(monkeypatch):
    """切れ目が 1 つも無い塊は、最後の手段として文字数で切る(拒否も切り詰めもしない)。"""
    monkeypatch.setattr(extract_xlsx, "MAX_CHUNK_CHARS", 50)
    chunks = extract_xlsx.extract("spec.xlsx", build({"S": [("A1", "あ" * 200)]}))
    assert [len(c["text"]) for c in chunks] == [50, 50, 50, 50]
    assert all(c["cells"] == "A1" for c in chunks)


def test_wide_row_over_limit_is_split_at_cell_boundaries(monkeypatch):
    """行が上限を超えても**各セルは収まる**なら、セル境界で割る(セルの中は切らない)。"""
    monkeypatch.setattr(extract_xlsx, "MAX_CHUNK_CHARS", 50)
    content = build({"S": [("A1", "あ" * 20), ("B1", "い" * 20), ("C1", "う" * 20)]})
    chunks = extract_xlsx.extract("wide.xlsx", content)
    assert [c["cells"] for c in chunks] == ["A1:B1", "C1"]
    assert [c["text"] for c in chunks] == ["あ" * 20 + "\t" + "い" * 20, "う" * 20]
    assert all("part" not in c for c in chunks)      # セル境界で足りた = 分割していない


def test_chunk_count_limit_still_applies_to_split_cells(monkeypatch):
    """セル内分割でも**総チャンク数の上限は据え置き**(超えたら従来どおり 422 相当)。"""
    monkeypatch.setattr(extract_xlsx, "MAX_CHUNKS", 3)
    with pytest.raises(extract_xlsx.ExtractionLimitError) as e:
        extract_xlsx.extract("spec.xlsx", build({"S": [("A1", _giant_cell_text(13_000))]}))
    assert e.value.limit == "chunks"


def test_long_block_is_split_at_row_boundary_not_truncated(monkeypatch):
    """上限に収まらない塊は**行境界で分割**する(本文は落とさない・出典は行範囲で残す)。"""
    monkeypatch.setattr(extract_xlsx, "MAX_CHUNK_CHARS", 50)
    rows = [(f"A{n}", "あ" * 20) for n in range(1, 7)]
    chunks = extract_xlsx.extract("long.xlsx", build({"S": rows}))
    assert len(chunks) == 3
    assert [c["cells"] for c in chunks] == ["A1:A2", "A3:A4", "A5:A6"]
    assert all(len(c["text"]) <= 50 for c in chunks)
    assert sum(c["text"].count("あ") for c in chunks) == 6 * 20   # 1 文字も落ちていない


def test_non_xlsx_bytes_are_rejected():
    with pytest.raises(extract_xlsx.UnsupportedWorkbook):
        extract_xlsx.extract("fake.xlsx", b"%PDF-1.7 not really a workbook")


# --- マネージド Vector Store 用の属性は「ファイル単位」 ------------------------


def test_file_attributes_are_file_level_for_multi_sheet_workbook():
    """複数シートのブックは、属性が**ファイル単位**であることを値として明示する。

    マネージド側の属性は 1 ファイルに 1 種類しか持てない(SPIKE-M1 ①-a)。
    チャンクの sheet / cells をそのまま載せると「セル単位で返る」という誤解を作るため、
    ブック全体を表す値にする(能力差を隠さない = ADR-0020 の決定内容)。
    """
    chunks = extract_xlsx.extract("spec.xlsx", build(SPEC))
    attrs = extract_xlsx.file_attributes(chunks)
    assert attrs == {"sheet": "(ブック全体: 2 シート)", "cells": "(ブック全体)"}
    assert attrs["cells"] not in {c["cells"] for c in chunks}


def test_file_attributes_use_bounding_range_for_single_sheet():
    content = build({"制約": [("C5", "レート制限"), ("E5", "600"), ("C9", "保持期間"),
                              ("E9", "13か月")]})
    chunks = extract_xlsx.extract("one.xlsx", content)
    assert len(chunks) == 2
    assert extract_xlsx.file_attributes(chunks) == {"sheet": "制約", "cells": "C5:E9"}


def test_file_attributes_of_empty_workbook_is_empty():
    assert extract_xlsx.file_attributes([]) == {}


def test_render_text_keeps_all_chunks():
    chunks = extract_xlsx.extract("spec.xlsx", build(SPEC))
    text = extract_xlsx.render_text(chunks)
    assert all(c["text"] in text for c in chunks)
    assert "[制約 C5:E5]" in text


# --- 壊れたブック・巨大なブックで 500 を漏らさない（review-2 PREP01-001/003） ---


def _corrupt_sheet_xml(content: bytes) -> bytes:
    """zip としては正しいが、シート XML が壊れているブックを作る。"""
    import zipfile

    out = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(content)) as src, zipfile.ZipFile(out, "w") as dst:
        for info in src.infolist():
            data = src.read(info.filename)
            if info.filename.startswith("xl/worksheets/"):
                data = b"<worksheet><sheetData><row r=\"1\"><c"   # 途中で切れた XML
            dst.writestr(info, data)
    return out.getvalue()


def test_broken_sheet_xml_is_rejected_not_raised_as_500():
    """`read_only` の反復は遅延実行なので、破損は行を読んでいる最中に初めて現れる。"""
    with pytest.raises(extract_xlsx.UnsupportedWorkbook):
        extract_xlsx.extract("corrupt.xlsx", _corrupt_sheet_xml(build(SPEC)))


def test_uncompressed_size_limit_is_rejected(monkeypatch):
    """圧縮率で殴られない（zip bomb）。展開後の合計を開く前に見る。"""
    monkeypatch.setattr(extract_xlsx, "MAX_UNCOMPRESSED_BYTES", 100)
    with pytest.raises(extract_xlsx.ExtractionLimitError) as e:
        extract_xlsx.extract("bomb.xlsx", build(SPEC))
    assert e.value.limit == "uncompressed_bytes"


def test_scan_budget_limit_is_rejected(monkeypatch):
    """チャンクにならないまま延々と読む形で止まる。"""
    monkeypatch.setattr(extract_xlsx, "MAX_SCANNED_CELLS", 3)
    with pytest.raises(extract_xlsx.ExtractionLimitError) as e:
        extract_xlsx.extract("many-rows.xlsx", build({"S": [(f"A{n}", f"値{n}")
                                                            for n in range(1, 30)]}))
    assert e.value.limit == "scanned_cells"


def test_scan_budget_counts_empty_rows(monkeypatch):
    """**空行も数える**。非空行だけ数えると、中身が無いのに広いブックが上限を抜ける。

    `read_only` の反復は欠けた行を空行で埋めるので、数 KB のブックでも
    「A1 と A50000 だけ」で 50,000 行の反復になる。
    """
    monkeypatch.setattr(extract_xlsx, "MAX_SCANNED_CELLS", 100)
    with pytest.raises(extract_xlsx.ExtractionLimitError) as e:
        extract_xlsx.extract("sparse.xlsx", build({"S": [("A1", "先頭"), ("A50000", "末尾")]}))
    assert e.value.limit == "scanned_cells"


def test_chunk_limit_stops_before_materialising_everything(monkeypatch):
    """上限は「1 チャンク作るたび」に見る（全部作ってから数えない）。

    1 つの塊が上限文字数で多数のチャンクに割れる形でも、`MAX_CHUNKS` を超えた時点で止まる。
    """
    monkeypatch.setattr(extract_xlsx, "MAX_CHUNK_CHARS", 30)
    monkeypatch.setattr(extract_xlsx, "MAX_CHUNKS", 3)
    rows = [(f"A{n}", "あ" * 20) for n in range(1, 60)]  # 連続する 1 塊 → 約 59 チャンク
    with pytest.raises(extract_xlsx.ExtractionLimitError) as e:
        extract_xlsx.extract("dense.xlsx", build({"S": rows}))
    assert e.value.limit == "chunks"


def test_row_that_extends_left_keeps_columns_aligned():
    """チャンクの途中で左に広がっても、列の対応が崩れない（描画幅の測り直し）。"""
    content = build({"S": [("C1", "右のみ"), ("A2", "左"), ("B2", "中"), ("C2", "右")]})
    chunks = extract_xlsx.extract("left.xlsx", content)
    assert len(chunks) == 1
    assert chunks[0]["cells"] == "A1:C2"
    assert chunks[0]["text"] == "\t\t右のみ\n左\t中\t右"


def test_widening_left_never_produces_an_oversized_chunk(monkeypatch):
    """左に広がった行が来ても、上限を超えるチャンクは作らない（行境界で確定する）。"""
    monkeypatch.setattr(extract_xlsx, "MAX_CHUNK_CHARS", 40)
    # C 列だけの行で上限近くまで埋め、次の行が A 列へ広がる（既存行の描画幅が 2 タブ増える）
    content = build({"S": [("C1", "あ" * 18), ("C2", "い" * 18),
                           ("A3", "左"), ("B3", "中"), ("C3", "右")]})
    chunks = extract_xlsx.extract("widen.xlsx", content)
    assert all(len(c["text"]) <= 40 for c in chunks)
    assert "".join(c["text"] for c in chunks).count("あ") == 18   # 本文は落ちていない


def _encrypted_zip(content: bytes) -> bytes:
    """暗号化フラグ（汎用目的ビット 0）を立てた zip を作る。

    zip としては開けるが、エントリを読む段で `zipfile` が **RuntimeError**（"password
    required"）を投げる。ValueError 系ではないので、正規化対象に入れ忘れると 500 で漏れる。
    フラグはヘッダのバイトを直接立てる（`writestr` はフラグを組み直すので効かない）。
    """
    raw = bytearray(content)
    # ローカルヘッダと中央ディレクトリの両方でフラグを立てる
    for signature, offset in ((b"PK\x03\x04", 6), (b"PK\x01\x02", 8)):
        at = raw.find(signature)
        while at != -1:
            raw[at + offset] |= 0x1
            at = raw.find(signature, at + 4)
    return bytes(raw)


def test_encrypted_workbook_is_rejected_not_raised_as_500():
    """パスワード付き（暗号化）ブックも、壊れたブックと同じ 422 相当で断る。"""
    with pytest.raises(extract_xlsx.UnsupportedWorkbook):
        extract_xlsx.extract("locked.xlsx", _encrypted_zip(build(SPEC)))
