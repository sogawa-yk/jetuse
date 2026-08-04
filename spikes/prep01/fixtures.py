"""PREP-01 の E2E で使う**架空**の仕様書ブック（顧客データは使わない）。

わざと扱いにくい要素を入れてある: 複数シート / 空シート / 結合セル /
キャッシュの無い数式セル / 大きな空白領域（行の飛び）。
v1 と v2 は「レート制限」の値だけが違い、版フィルタの対照に使える。
"""

import io

from openpyxl import Workbook

# OCI 側・検証用の資源には接頭辞を付ける（CLAUDE.md の運用規約）
PREFIX = "jetuse-spike-prep01"
DOC_NAME = f"{PREFIX}-サンプル在庫連携API仕様書.xlsx"

RATE_LIMIT = {"1.0": "300 req/min", "2.0": "600 req/min"}


def workbook(version: str) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    api = wb.create_sheet("API一覧")
    api["B12"] = "エンドポイント"
    api["C12"] = "メソッド"
    api["D12"] = "説明"
    api["B13"] = "/v1/inventory"
    api["C13"] = "GET"
    api["D13"] = "在庫数と引当可能数を返す在庫照会API"
    api["B14"] = "/v1/shipments"
    api["C14"] = "POST"
    api["D14"] = "出荷指示を登録する"

    seigen = wb.create_sheet("制約")
    seigen["A1"] = "本仕様書の制約事項"          # 結合セルの見出し
    seigen.merge_cells("A1:E1")
    seigen["C5"] = "レート制限"
    seigen["D5"] = RATE_LIMIT[version]
    seigen["E5"] = "超過時は HTTP 429 を返す"
    seigen["C6"] = "データ保持期間"
    seigen["D6"] = "13か月"
    seigen["E6"] = "明細データが対象"
    # 大きな空白領域を挟んだ別の塊（別チャンクになること）
    seigen["C40"] = "同時接続数"
    seigen["D40"] = "50"
    seigen["E40"] = "IP 単位で計数する"
    seigen["D41"] = "=SUM(D40:D40)"            # 値のキャッシュが無い数式セル

    rev = wb.create_sheet("改訂履歴")
    rev["A1"] = "版"
    rev["B1"] = "日付"
    rev["C1"] = "内容"
    rev["A2"] = version
    rev["B2"] = "2026-07-30"
    rev["C2"] = f"レート制限を {RATE_LIMIT[version]} に改訂"

    wb.create_sheet("作業用")                   # 空シート（チャンクを作らないこと）

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
