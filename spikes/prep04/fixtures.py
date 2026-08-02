"""PREP-04 の E2E で使う**架空**のブック（顧客の実ファイルは使わない・持ち込まない）。

再現したいのは実データの**形**だけ:「1 つのセルに API のリクエスト例が丸ごと入っていて、
その行の非空セルは 1 個」（実データは A53 に 13,728 文字）。中身は架空の在庫連携 API。

- `giant_workbook()`: `A53` に 13,000 文字級の 1 セル（= 行境界でもセル境界でも割れない）
- `plain_workbook()`: 上限に掛からない通常のブック（回帰の対照。PREP-01 と同じ形）
"""

import io

from openpyxl import Workbook

# OCI 側・検証用の資源には接頭辞を付ける（CLAUDE.md の運用規約）
PREFIX = "jetuse-spike-prep04"
GIANT_NAME = f"{PREFIX}-サンプル在庫連携API仕様書-リクエスト例.xlsx"
PLAIN_NAME = f"{PREFIX}-サンプル在庫連携API仕様書-通常.xlsx"

GIANT_CELL = "A53"          # 実データと同じ位置（1 行 = 1 セル）
GIANT_SHEET = "サンプル"

# 巨大セルの**中ほど**に置く一文。ここがヒットすれば「先頭 2,000 文字だけが検索対象」
# ではないことが示せる（分割前は先頭で切られるか、そもそも取り込めなかった）。
MARKER = "冪等キー(idempotency_key)は発行から24時間有効"
GIANT_QUESTION = "冪等キーは何時間有効ですか"

RATE_LIMIT = "600 req/min"
PLAIN_QUESTION = "レート制限は1分あたり何リクエストですか"


# 1 セルの中は**話題ごとに違う内容**にする（実物の仕様書と同じ）。全部が同じ形の
# JSON だと、どの断片も同じように埋め込まれて検索が話題を区別できず、
# 「セルの中ほどがヒットするか」を測れない（実装ではなく検証データの問題）。
_TOPICS = [
    ("在庫同期", "/v1/inventory/sync", "warehouse", "在庫数と引当可能数を更新する",
     "在庫が負になる更新は 409 を返し、部分適用はしない"),
    ("出荷指示", "/v1/shipments", "carrier", "出荷指示を登録し追跡番号を採番する",
     "出荷済みの指示は取り消せない。訂正は返品伝票で行う"),
    ("返品受領", "/v1/returns", "reason_code", "返品の受領を記録して在庫へ戻す",
     "検品が未了の返品は引当対象に入れない"),
    ("入荷予定", "/v1/purchase-orders", "supplier", "入荷予定を登録する",
     "入荷予定日を過去日にすると 422 を返す"),
    ("棚卸差異", "/v1/stocktakes", "location", "棚卸の差異を登録する",
     "差異が閾値を超えると承認待ちになる"),
]


def _section(n: int) -> str:
    """架空の API 仕様の 1 節（説明 + リクエスト例 + 注意）。改行 = 意味の切れ目。"""
    name, path, key, summary, caution = _TOPICS[n % len(_TOPICS)]
    return (
        f"### {n}. {name} API\n"
        f"{summary}。呼び出しは `POST {path}` で、`{key}` は必須項目である。\n"
        f"注意: {caution}。\n"
        "リクエスト例:\n"
        "{\n"
        f'  "request_id": "REQ-{n:04d}",\n'
        f'  "{key}": "{key.upper()[:2]}-{n:03d}",\n'
        '  "items": [\n'
        f'    {{"sku": "SKU-{n:04d}", "qty": {n % 50 + 1}, "lot": "L-2026-{n:04d}"}}\n'
        "  ]\n"
        "}\n"
        f"レスポンス例: {{\"status\": \"accepted\", \"request_id\": \"REQ-{n:04d}\"}}\n"
    )


def _idempotency_section() -> str:
    """セルの中ほどに置く節（ここだけが冪等性の話をしている = 検索で区別できる）。"""
    return (
        "### 冪等性の扱い\n"
        f"{MARKER}。有効期間内に同一の冪等キーで再送された要求は、最初の結果をそのまま返す。\n"
        "同一キーの再送に対しては 200 を返し、在庫を二重に更新しない。\n"
        "有効期間を過ぎた冪等キーは失効し、同じキーでも新しい要求として処理される。\n"
        "冪等キーを付けずに送った要求は再送のたびに別の要求として扱われるので、"
        "ネットワーク再送のある経路では必ず付けること。\n"
    )


def giant_cell_text(target: int = 13_000) -> str:
    """1 セルに入れる 13,000 文字級のテキスト（末尾は節の区切りで終わる）。"""
    blocks: list[str] = []
    total = 0
    n = 1
    inserted = False
    while total < target:
        if not inserted and total >= target * 0.5:
            block = _idempotency_section()
            inserted = True
        else:
            block = _section(n)
            n += 1
        blocks.append(block)
        total += len(block)
    return "".join(blocks).strip()


def giant_workbook() -> bytes:
    """実データと同じ形: 1 行の非空セルが 1 個で、そのセルが上限を大きく超える。"""
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(GIANT_SHEET)
    ws["A1"] = "在庫連携API リクエスト例（サンプル）"
    ws["A2"] = "以下に代表的なリクエストとレスポンスの例を示す。"
    ws[GIANT_CELL] = giant_cell_text()      # ここだけで 13,000 文字級
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def plain_workbook() -> bytes:
    """回帰の対照。上限に掛からない通常の表（PREP-01 が固定した形）。"""
    wb = Workbook()
    wb.remove(wb.active)

    api = wb.create_sheet("API一覧")
    api["B12"] = "エンドポイント"
    api["C12"] = "メソッド"
    api["D12"] = "説明"
    api["B13"] = "/v1/inventory"
    api["C13"] = "GET"
    api["D13"] = "在庫数と引当可能数を返す在庫照会API"

    seigen = wb.create_sheet("制約")
    seigen["C5"] = "レート制限"
    seigen["D5"] = RATE_LIMIT
    seigen["E5"] = "超過時は HTTP 429 を返す"
    seigen["C40"] = "同時接続数"
    seigen["D40"] = "50"
    seigen["E40"] = "IP 単位で計数する"

    wb.create_sheet("作業用")               # 空シート（チャンクを作らない）

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
